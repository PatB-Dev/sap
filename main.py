import win32com.client
import easygui
from openpyxl.reader.excel import load_workbook
import os
from sapconnect import get_client

file_data = load_workbook('python.xlsx')
sheet = file_data["Feuil1"]

lineDebut = int(easygui.enterbox('N° de ligne début'))
lineFin = int(easygui.enterbox('N° de ligne fin'))

while lineDebut <= lineFin:
    def main():
        # Initialize an object to store the SAP session
        obj_sess = get_client()

        # Replace 'session' with initialized 'obj_sess'
        obj_sess.findById("wnd[0]/tbar[0]/okcd").text = "MM03"
        obj_sess.findById("wnd[0]").sendVKey(0)
        obj_sess.findById("wnd[0]/usr/ctxtRMMG1-MATNR").text = sheet.cell(lineDebut, 1).value
        obj_sess.findById("wnd[0]/usr/ctxtRMMG1-MATNR").caretPosition = 17
        obj_sess.findById("wnd[0]").sendVKey(0)
        obj_sess.findById("wnd[1]/tbar[0]/btn[0]").press()
        obj_sess.findById("wnd[1]/tbar[0]/btn[0]").press()
        designation = obj_sess.findById("wnd[0]/usr/subSUB2:SAPLMGD1:8001/tblSAPLMGD1TC_KTXT/txtSKTEXT-MAKTX[1,0]").Text
        obj_sess.findById("wnd[0]/usr/subSUB2:SAPLMGD1:8001/tblSAPLMGD1TC_KTXT/txtSKTEXT-MAKTX[1,0]").caretPosition = 0
        obj_sess.findById("wnd[0]/tbar[0]/btn[3]").press()
        obj_sess.findById("wnd[0]/tbar[0]/btn[3]").press()
        sheet.cell(row=lineDebut, column=2).value = designation
        # array = [designation]
        # df = pandas.DataFrame(array)
        # df.to_excel('24082023.xlsx', sheet_name='patTest')

    if __name__ == '__main__':
        main()
    lineDebut += 1
file_data.save('python.xlsx')
os.system('python.xlsx')
